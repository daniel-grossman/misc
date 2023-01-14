#adds (and alphabetically sorts) words + their definitions to a spreadsheet
#uses Dictionary.com API, you will need API key (it's free)

import openpyxl
import requests
import json

spreadsheet = '~/PATHWAY-TO-EXCEL-FILE'
workbook = openpyxl.load_workbook(spreadsheet)
worksheet = workbook.active

word_list = []
#enter all the new words you want added to your dictionary spreadsheet
while True:
    word = str(input('Enter word: '))
    if word == '':
        break
    word_list.append(word)

#gets each word's definition and adds both to spreadsheet
for word in word_list:
    url = 'https://dictionaryapi.com/api/v3/references/collegiate/json/' + word + 'YOUR-DICTIONARY.COM-API-KEY'
    response = requests.get(url)
    data = json.loads(response.text)
    definition = data[0]['shortdef'][0]
    print(word + ': ' + definition)
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1).value = word
    worksheet.cell(row=next_row, column=2).value = definition
    workbook.save(spreadsheet)

#sorts ALL the words in the spreadsheet
word_data = [cell.value for cell in worksheet['A']]
definition_data = [cell.value for cell in worksheet['B']]
sorted_data = list(zip(word_data, definition_data))
sorted_data.sort(key=lambda x: x[0])

for i, value in enumerate(sorted_data, 1):
    worksheet.cell(row=i,column=1).value = value[0]
    worksheet.cell(row=i,column=2).value = value[1]

workbook.save(spreadsheet)
